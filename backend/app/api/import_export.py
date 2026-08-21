"""
数据导入导出 API
支持 Excel/Xmind 格式的测试用例导入导出
"""
import json
import uuid
import zipfile
from datetime import datetime
from app.core.timezone import china_now_naive
from typing import List
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from app.database import get_db
from app.core.deps import get_current_user, get_project
from app.core.audit import log_audit
from app.models.user import User
from app.models.test_case import TestCase
from app.models.requirement import TestRequirement

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter()
project_router = APIRouter(prefix="/api/projects/{project_id}/data")


def _steps_to_text(steps_raw: str) -> str:
    """将测试步骤 JSON 转为自然语言文本"""
    if not steps_raw:
        return ""
    try:
        steps_list = json.loads(steps_raw)
    except (json.JSONDecodeError, TypeError):
        return steps_raw

    parts = []
    for i, step in enumerate(steps_list):
        if isinstance(step, dict):
            action = step.get("action", "")
            expected = step.get("expected", "")
            parts.append(f"步骤{i+1}：{action}，预期结果：{expected}")
        elif isinstance(step, str):
            parts.append(f"步骤{i+1}：{step}")
    return "；".join(parts)


# Excel 列定义
CASE_COLUMNS = [
    ("用例标题", "title", 30),
    ("所属模块", "module", 20),
    ("优先级", "priority", 10),
    ("用例类型", "case_type", 15),
    ("前置条件", "preconditions", 30),
    ("测试步骤", "steps", 50),
    ("预期结果", "expected_result", 40),
    ("状态", "status", 10),
    ("关联需求ID", "req_id", 15),
    ("BDD内容", "bdd_content", 40),
]


@project_router.get("/cases/export")
def export_cases(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出测试用例为 Excel"""
    get_project(project_id, db, current_user)
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请运行 pip install openpyxl")

    cases = db.query(TestCase).filter(
        TestCase.project_id == project_id,
        TestCase.is_deleted == False,
    ).order_by(TestCase.module, TestCase.priority).all()

    log_audit(
        db, action="export", resource_type="case",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "count": len(cases), "format": "excel"},
    )
    db.commit()

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 写入表头
    for col_idx, (header, _, width) in enumerate(CASE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = width

    # 写入数据
    for row_idx, case in enumerate(cases, 2):
        for col_idx, (_, field, _) in enumerate(CASE_COLUMNS, 1):
            value = getattr(case, field, "")
            if field == "steps":
                value = _steps_to_text(value) if isinstance(value, str) else value
            ws.cell(row=row_idx, column=col_idx, value=value or "")

    # 生成文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"test_cases_{project_id}_{china_now_naive().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@project_router.post("/cases/import")
async def import_cases(
    project_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """从 Excel 导入测试用例"""
    get_project(project_id, db, current_user)
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请运行 pip install openpyxl")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件格式 (.xlsx, .xls)")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {str(e)}")

    # 读取表头，建立列映射
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            headers[str(header_value).strip()] = col

    # 字段映射
    field_map = {}
    for header_name, field_name, _ in CASE_COLUMNS:
        if header_name in headers:
            field_map[field_name] = headers[header_name]

    if "title" not in field_map:
        raise HTTPException(status_code=400, detail="Excel 必须包含「用例标题」列")

    # 读取数据行
    imported = 0
    failed = 0
    errors = []

    for row in range(2, ws.max_row + 1):
        try:
            title = ws.cell(row=row, column=field_map.get("title", 0)).value
            if not title:
                continue

            case_data = {
                "project_id": project_id,
                "title": str(title).strip(),
                "created_by": current_user.id,
            }

            for field_name, col_idx in field_map.items():
                if field_name == "title":
                    continue
                value = ws.cell(row=row, column=col_idx).value
                if value is not None and value != "":
                    if field_name == "steps":
                        # 尝试解析步骤
                        steps_text = str(value)
                        steps = []
                        for line in steps_text.split("\n"):
                            line = line.strip()
                            if line:
                                # 移除序号前缀
                                import re
                                line = re.sub(r'^\d+[\.\、]\s*', '', line)
                                steps.append(line)
                        case_data["steps"] = json.dumps(steps, ensure_ascii=False) if steps else json.dumps([steps_text], ensure_ascii=False)
                    elif field_name == "req_id":
                        try:
                            case_data["req_id"] = int(value)
                        except (ValueError, TypeError):
                            pass
                    else:
                        case_data[field_name] = str(value).strip()

            # 设置默认值
            case_data.setdefault("module", "默认模块")
            case_data.setdefault("priority", "P2")
            case_data.setdefault("case_type", "functional")
            case_data.setdefault("status", "active")
            case_data.setdefault("preconditions", "")
            case_data.setdefault("expected_result", "")
            case_data.setdefault("steps", json.dumps(["执行测试"], ensure_ascii=False))

            case = TestCase(**case_data)
            db.add(case)
            imported += 1

        except Exception as e:
            failed += 1
            errors.append(f"第 {row} 行: {str(e)}")

    log_audit(
        db, action="import", resource_type="case",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "filename": file.filename, "imported": imported, "failed": failed},
        status="success" if failed == 0 else "partial",
    )
    db.commit()

    return {
        "detail": "导入完成",
        "imported": imported,
        "failed": failed,
        "errors": errors[:10]  # 最多返回前10个错误
    }


@project_router.get("/cases/template")
def download_template(project_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """下载用例导入模板"""
    get_project(project_id, db, current_user)
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请运行 pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例模板"

    # 写入表头
    for col_idx, (header, _, width) in enumerate(CASE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = width

    # 写入示例数据
    example = {
        "title": "用户登录成功",
        "module": "用户管理",
        "priority": "P0",
        "case_type": "functional",
        "preconditions": "用户已注册",
        "steps": "1. 打开登录页\n2. 输入用户名密码\n3. 点击登录",
        "expected_result": "登录成功，跳转到首页",
        "status": "active",
        "req_id": "",
        "bdd_content": "",
    }
    for col_idx, (_, field, _) in enumerate(CASE_COLUMNS, 1):
        ws.cell(row=2, column=col_idx, value=example.get(field, ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_case_template.xlsx"}
    )


def _make_xmind_topic(title: str, children: list = None) -> dict:
    """构建 XMind topic 节点"""
    topic = {"id": uuid.uuid4().hex, "class": "topic", "title": title}
    if children:
        topic["children"] = {"attached": children}
    return topic


@project_router.get("/cases/export-xmind")
def export_cases_xmind(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出测试用例为 XMind 文件，节点结构：所属模块 → 用例标题 → 前置条件 → 测试步骤 → 预期结果"""
    get_project(project_id, db, current_user)

    cases = db.query(TestCase).filter(
        TestCase.project_id == project_id,
        TestCase.is_deleted == False,
    ).order_by(TestCase.module, TestCase.priority).all()

    log_audit(
        db, action="export", resource_type="case",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "count": len(cases), "format": "xmind"},
    )
    db.commit()

    # 按模块分组
    modules: dict[str, list] = {}
    for case in cases:
        module_name = case.module or "未分类模块"
        modules.setdefault(module_name, []).append(case)

    # 构建 XMind 树
    module_topics = []
    for module_name, module_cases in modules.items():
        case_topics = []
        for case in module_cases:
            steps_text = _steps_to_text(case.steps) if case.steps else "无步骤"
            expected_topic = _make_xmind_topic(f"预期结果：{case.expected_result or '无'}")
            steps_topic = _make_xmind_topic(f"测试步骤：{steps_text}", [expected_topic])
            precond_topic = _make_xmind_topic(f"前置条件：{case.preconditions or '无'}", [steps_topic])
            case_topics.append(_make_xmind_topic(case.title, [precond_topic]))

        module_topics.append(_make_xmind_topic(module_name, case_topics))

    root_topic = _make_xmind_topic("测试用例", module_topics)

    content = [{
        "id": uuid.uuid4().hex,
        "class": "sheet",
        "title": "测试用例",
        "rootTopic": root_topic,
    }]

    metadata = {"creator": {"name": "AITS", "version": "1.0"}}

    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
        }
    }

    # 生成 XMind 文件（ZIP 格式）
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False))
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
    output.seek(0)

    filename = f"test_cases_{project_id}_{china_now_naive().strftime('%Y%m%d_%H%M%S')}.xmind"

    return StreamingResponse(
        output,
        media_type="application/vnd.xmind.workbook",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== P2-12: API 测试数据导出/导入 ====================


def _serialize_model(obj, exclude_fields=None):
    """序列化 SQLAlchemy 模型为字典"""
    exclude_fields = exclude_fields or set()
    data = {}
    for col in obj.__table__.columns:
        if col.name in exclude_fields:
            continue
        val = getattr(obj, col.name)
        if hasattr(val, "isoformat"):
            data[col.name] = val.isoformat()
        else:
            data[col.name] = val
    return data


@project_router.get("/api-definitions/export")
def export_api_definitions(
    project_id: int,
    request: Request,
    module_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出 API 接口定义为 JSON"""
    from app.models.api_test import ApiDefinition, ApiModule

    query = db.query(ApiDefinition).filter(
        ApiDefinition.project_id == project_id,
        ApiDefinition.is_deleted == False,
    )
    if module_id:
        query = query.filter(ApiDefinition.module_id == module_id)

    definitions = query.order_by(ApiDefinition.method, ApiDefinition.path).all()
    modules = db.query(ApiModule).filter(
        ApiModule.project_id == project_id,
        ApiModule.is_deleted == False,
    ).all()

    data = {
        "export_type": "api_definitions",
        "version": "1.0",
        "project_id": project_id,
        "exported_at": china_now_naive().isoformat(),
        "modules": [_serialize_model(m, {"is_deleted", "deleted_at"}) for m in modules],
        "definitions": [_serialize_model(d, {"is_deleted", "deleted_at"}) for d in definitions],
    }

    log_audit(
        db, action="export", resource_type="api_definition",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "count": len(definitions), "format": "json"},
    )
    db.commit()

    filename = f"api_definitions_{project_id}_{china_now_naive().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@project_router.post("/api-definitions/import")
async def import_api_definitions(
    project_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从 JSON 导入 API 接口定义"""
    from app.models.api_test import ApiDefinition, ApiModule

    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"JSON 解析失败: {str(e)}")

    if data.get("export_type") != "api_definitions":
        raise HTTPException(status_code=400, detail="文件类型不匹配，期望 api_definitions")

    imported_modules = 0
    imported_defs = 0
    module_id_map = {}

    for mod_data in data.get("modules", []):
        old_id = mod_data.pop("id", None)
        mod = ApiModule(
            **mod_data,
            project_id=project_id,
        )
        db.add(mod)
        db.flush()
        if old_id:
            module_id_map[old_id] = mod.id
        imported_modules += 1

    for def_data in data.get("definitions", []):
        old_module_id = def_data.pop("module_id", None)
        if old_module_id and old_module_id in module_id_map:
            def_data["module_id"] = module_id_map[old_module_id]
        def_data.pop("id", None)
        def_data["project_id"] = project_id
        def_obj = ApiDefinition(**def_data)
        db.add(def_obj)
        imported_defs += 1

    log_audit(
        db, action="import", resource_type="api_definition",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "modules": imported_modules, "definitions": imported_defs},
    )
    db.commit()

    return {
        "detail": "导入完成",
        "imported_modules": imported_modules,
        "imported_definitions": imported_defs,
    }


@project_router.get("/api-cases/export")
def export_api_cases(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出接口测试用例（含断言）为 JSON"""
    from app.models.api_test import ApiTestCase, ApiCaseAssertion

    cases = db.query(ApiTestCase).filter(
        ApiTestCase.project_id == project_id,
        ApiTestCase.is_deleted == False,
    ).order_by(ApiTestCase.created_at.desc()).all()

    case_data_list = []
    for case in cases:
        case_dict = _serialize_model(case, {"is_deleted", "deleted_at"})
        assertions = db.query(ApiCaseAssertion).filter(
            ApiCaseAssertion.case_id == case.id,
            ApiCaseAssertion.is_deleted == False,
        ).all()
        case_dict["assertions"] = [_serialize_model(a, {"is_deleted", "deleted_at", "case_id"}) for a in assertions]
        case_data_list.append(case_dict)

    data = {
        "export_type": "api_cases",
        "version": "1.0",
        "project_id": project_id,
        "exported_at": china_now_naive().isoformat(),
        "cases": case_data_list,
    }

    log_audit(
        db, action="export", resource_type="api_case",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "count": len(cases), "format": "json"},
    )
    db.commit()

    filename = f"api_cases_{project_id}_{china_now_naive().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@project_router.post("/api-cases/import")
async def import_api_cases(
    project_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从 JSON 导入接口测试用例"""
    from app.models.api_test import ApiTestCase, ApiCaseAssertion

    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"JSON 解析失败: {str(e)}")

    if data.get("export_type") != "api_cases":
        raise HTTPException(status_code=400, detail="文件类型不匹配，期望 api_cases")

    imported_cases = 0
    imported_assertions = 0

    for case_data in data.get("cases", []):
        assertions = case_data.pop("assertions", [])
        case_data.pop("id", None)
        case_data["project_id"] = project_id
        case_obj = ApiTestCase(**case_data)
        db.add(case_obj)
        db.flush()

        for assert_data in assertions:
            assert_data.pop("id", None)
            assert_data["case_id"] = case_obj.id
            assert_obj = ApiCaseAssertion(**assert_data)
            db.add(assert_obj)
            imported_assertions += 1

        imported_cases += 1

    log_audit(
        db, action="import", resource_type="api_case",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "cases": imported_cases, "assertions": imported_assertions},
    )
    db.commit()

    return {
        "detail": "导入完成",
        "imported_cases": imported_cases,
        "imported_assertions": imported_assertions,
    }


@project_router.get("/api-scenarios/export")
def export_api_scenarios(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出接口测试场景（含步骤和变量）为 JSON"""
    from app.models.api_test import ApiScenario, ApiScenarioStep, ApiScenarioVariable

    scenarios = db.query(ApiScenario).filter(
        ApiScenario.project_id == project_id,
        ApiScenario.is_deleted == False,
    ).order_by(ApiScenario.created_at.desc()).all()

    scenario_list = []
    for sc in scenarios:
        sc_dict = _serialize_model(sc, {"is_deleted", "deleted_at"})
        steps = db.query(ApiScenarioStep).filter(
            ApiScenarioStep.scenario_id == sc.id,
            ApiScenarioStep.is_deleted == False,
        ).order_by(ApiScenarioStep.sort_order).all()
        sc_dict["steps"] = [_serialize_model(s, {"is_deleted", "deleted_at", "scenario_id"}) for s in steps]
        variables = db.query(ApiScenarioVariable).filter(
            ApiScenarioVariable.scenario_id == sc.id,
            ApiScenarioVariable.is_deleted == False,
        ).all()
        sc_dict["variables"] = [_serialize_model(v, {"is_deleted", "deleted_at", "scenario_id"}) for v in variables]
        scenario_list.append(sc_dict)

    data = {
        "export_type": "api_scenarios",
        "version": "1.0",
        "project_id": project_id,
        "exported_at": china_now_naive().isoformat(),
        "scenarios": scenario_list,
    }

    log_audit(
        db, action="export", resource_type="api_scenario",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "count": len(scenarios), "format": "json"},
    )
    db.commit()

    filename = f"api_scenarios_{project_id}_{china_now_naive().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@project_router.post("/api-scenarios/import")
async def import_api_scenarios(
    project_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从 JSON 导入接口测试场景"""
    from app.models.api_test import ApiScenario, ApiScenarioStep, ApiScenarioVariable

    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"JSON 解析失败: {str(e)}")

    if data.get("export_type") != "api_scenarios":
        raise HTTPException(status_code=400, detail="文件类型不匹配，期望 api_scenarios")

    imported_scenarios = 0
    imported_steps = 0
    imported_variables = 0

    for sc_data in data.get("scenarios", []):
        steps = sc_data.pop("steps", [])
        variables = sc_data.pop("variables", [])
        sc_data.pop("id", None)
        sc_data["project_id"] = project_id
        sc_obj = ApiScenario(**sc_data)
        db.add(sc_obj)
        db.flush()

        for step_data in steps:
            step_data.pop("id", None)
            step_data["scenario_id"] = sc_obj.id
            step_obj = ApiScenarioStep(**step_data)
            db.add(step_obj)
            imported_steps += 1

        for var_data in variables:
            var_data.pop("id", None)
            var_data["scenario_id"] = sc_obj.id
            var_obj = ApiScenarioVariable(**var_data)
            db.add(var_obj)
            imported_variables += 1

        imported_scenarios += 1

    log_audit(
        db, action="import", resource_type="api_scenario",
        user=current_user,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
        detail={"project_id": project_id, "scenarios": imported_scenarios, "steps": imported_steps, "variables": imported_variables},
    )
    db.commit()

    return {
        "detail": "导入完成",
        "imported_scenarios": imported_scenarios,
        "imported_steps": imported_steps,
        "imported_variables": imported_variables,
    }
